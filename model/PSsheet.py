import openpyxl as xl
import xlwings as xw
from Worksheet import Worksheet,QPreviewItem
from Workcell import *
from PyQt4.QtGui import *
from PyQt4.QtCore import *
from copy import copy
#import time
import datetime
##################################################
#       class for PS sheet handling
##################################################
UP = 0
DOWN = 1
nn = 0
class PSsheet(Worksheet):
    def __init__(self,sheet = None,sheet_wr = None):
        super(PSsheet,self).__init__(sheet,sheet_wr)
        self._preview_model_list = []
        self._extended_preview_model_list = None
        self.init_ps_sheet()
        self.init_ps_model()
    def __del__(self):
        super(PSsheet,self).__del__()
        del self._preview_model
        del self._extended_preview_model_list
    
    def init_ps_sheet(self):
        if self._xmlname != None:
            self._status = self.search_header_by_value(u'Status(POR,INIT,PREV)')
            self._subject_matter = self.search_header_by_value(u'Subject Matter/\nFunctional Area')
            self._container_name = self.search_header_by_value(u'Container Name\nTechnical Specification')
    def init_ps_model(self):
        self._preview_model = QStandardItemModel()
        self._preview_model.setColumnCount(4)
        self._preview_model.setHeaderData(0,Qt.Horizontal,'status')
        self._preview_model.setHeaderData(1,Qt.Horizontal,'subject matter')
        self._preview_model.setHeaderData(2,Qt.Horizontal,'container name')
        self._preview_model.setHeaderData(3,Qt.Horizontal,'xmlname')
        self._extended_preview_model = None
    def update_model(self):
        super(PSsheet,self).update_model()
        self._preview_model_list = []
        self.init_ps_model()
        self.init_ps_sheet()
        if self._xmlname != None:
            try:
                for xml_name in self.xml_names():
                    item_status = QPreviewItem(self._status.get_item_by_xmlname(xml_name))
                    item_subject_matter = QPreviewItem(self._subject_matter.get_item_by_xmlname(xml_name))
                    item_container_name = QPreviewItem(self._container_name.get_item_by_xmlname(xml_name))
                    item_xml_name = QPreviewItem(xml_name)
                    self._preview_model.appendRow((item_status,item_subject_matter,item_container_name,item_xml_name))
                    self._preview_model_list.append((item_status.value,item_subject_matter.value,item_container_name.value,item_xml_name.value))
            except:
                return 'error'
    def status(self):
        cells = list(self._worksheet.iter_cols(min_col=self._status.col,min_row=self._status.row+1,max_col=self._status.col,max_row=self.max_row).next())
        while cells[-1].value == None:
            cells.pop()
        return map(lambda x:Status(x,self._worksheet_wr),cells)
    def cell(self,row,col):
        #return self._worksheet_wr.range(row,col)
        return self._worksheet.cell(row=row,column=col)
    
    def auto_fit(self,cols):
        for col in cols:
            for i in range(1):
                self._worksheet_wr.api.Columns(col).AutoFit()

    def add_row(self,start_pos,offset,orientation):
        loop = offset
        while loop > 0:
            self._worksheet_wr.api.Rows[start_pos].Insert(-4121)
            loop -= 1


    def delete_row(self,start_pos,offset):
        self._worksheet_wr.range('%d:%d'%(start_pos,start_pos+offset-1)).api.Delete()
    def lock_row(self,row,status):
        self._worksheet_wr.api.Rows[row-1].Locked = True
    def lock_sheet(self):
        self._worksheet_wr.api.Protect()
    def lock_sheet_status(self):
        return self._worksheet_wr.api.ProtectContents
    def unlock_sheet(self):
        self._worksheet_wr.api.Unprotect()
    def unlock_all_cells(self):
        self._worksheet_wr.api.Cells.Locked = False

    def extended_preview_model(self):
        if self._extended_preview_model == None:
            self._extended_preview_model = QStandardItemModel()
            self._extended_preview_model.setColumnCount(self._worksheet.max_column)
            for row in self._worksheet.rows:
                item_row = []
                for cell in row:
                    try:
                        if cell.value == None:
                            item = QStandardItem('')
                        else:
                            item = QStandardItem(cell.value)
                    except:
                        item = QStandardItem('')
                    item_row.append(item)
                self._extended_preview_model.appendRow(item_row)
        return self._extended_preview_model
    @property
    def extended_preview_model_list(self):
        if self._extended_preview_model_list == None:
            self._extended_preview_model_list = []
            for row in self._worksheet.iter_rows(min_row=self.min_row,max_row=self.max_row,min_col=self.min_col,max_col=self.max_col):
                item_row = []
                for cell in row:
                    #item_row.append(cell.value if cell.value is not None else '')
                    try:
                        #item_row.append('' if cell.value is None else str(cell.value) if type(cell.value) == long else '%s-%s-%s'%((cell+datetime.delta(days=1)).timetuple().tm_year,(cell+datetime.delta(days=1)).timetuple().tm_mon,(cell+datetime.delta(days=1)).timetuple().tm_mday) if type(cell) is datetime.datetime else (cell.value))
                        item_row.append('' if cell.value is None else str('%s-%s-%s'%((cell.value+datetime.timedelta(days=1)).timetuple().tm_year,(cell.value+datetime.timedelta(days=1)).timetuple().tm_mon,(cell.value+datetime.timedelta(days=1)).timetuple().tm_mday)) if type(cell.value) is datetime.datetime else str(cell.value))
                    except:
                        item_row.append('')
                self._extended_preview_model_list.append(item_row)
        return self._extended_preview_model_list
        
    @property
    def preview_model(self):
        return self._preview_model_list
