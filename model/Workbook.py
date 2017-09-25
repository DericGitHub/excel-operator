import openpyxl as xl
import xlwings as xw

from Worksheet import Worksheet
from PyQt4.QtGui import *
from PyQt4.QtCore import *
from io import BytesIO,StringIO

##################################################
#       abstract class to handle workbook
##################################################
class Workbook(object):

    ##################################################
    #       Initial method
    ##################################################
    def __init__(self,workbook = None,app = None):
        self._workbook_wr = None
        self._sheets = []
        if workbook != None:
            self.init_book(workbook,app)
        self.init_model()
    def __del__(self):
        del self._sheets
        if self._workbook_wr is not None:
            print 'close wr book %s'%self._workbook_wr
            self._workbook_wr.close()

    def init_book(self,workbook,app):
        if isinstance(workbook,BytesIO):
            self._workbook_name = workbook
        else:
            self._workbook_name = str(workbook)
        self._sheets_cnt = None
        self._sheets_name = []
        self._workbook = xl.load_workbook(self._workbook_name)
        if app == None:
            self._workbook_wr = xw.Book(self._workbook_name)
        else:
            self._workbook_wr = app.books.open(self._workbook_name)
        print 'open wr book %s'%self._workbook_wr
        self._current_sheet = None
    def init_model(self):
        self._book_name = None
        self._sheet_name_model = QStandardItemModel()
    def update_model(self):
        self.update_sheet_name_model()
    def update_sheet_name_model(self):
        self._sheet_name_model.clear()
        for sheetname in self._workbook.sheetnames:
            item_sheet_name = QStandardItem(sheetname)
            self._sheet_name_model.appendRow(item_sheet_name)
            

    def load_sheets(self,sheet_cls,sheets,sheets_wr):
        for i in range(len(sheets)):
            self._sheets.append(sheet_cls(sheets[i],sheets_wr[i]))

    def load_sheets_name(self,sheets):
        for sheet in sheets:
            self._sheets_name.append(sheet)

    ##################################################
    #       user interface
    ##################################################
    def open(self,path_name):
        pass
    def save_as(self,path_name):
        self._workbook_wr.save(path_name)
    def save(self):
        pass
    def recover(self,direction):
        pass

    ##################################################
    #       data 
    ##################################################
    @property
    def sheets(self):
        return self._sheets
    @property
    def sheet_name_model(self):
        return self._sheet_name_model
    @property
    def workbook(self):
        return self._workbook
    @property
    def workbook_wr(self):
        return self._workbook_wr
    @property
    def workbook_name(self):
        return self._workbook_name
    @property
    def sheetnames(self):
        return self._workbook.sheetnames
