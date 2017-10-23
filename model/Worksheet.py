import openpyxl as xl
import xlwings as xw
from Workcell import *
from PyQt4.QtGui import *
from PyQt4.QtCore import *
from configparser import ConfigParser as cp
try:
    cfg = cp()
    cfg.read('setup.ini')
    MAX_ROW = cfg.getint('setup','MAX_ROW')
except:
    MAX_ROW = 1000

##################################################
#       Abstract class to handle worksheet
##################################################
class Worksheet(object):

    ##################################################
    #       Initial method
    ##################################################
    def __init__(self,sheet = None,sheet_wr = None):
        self._worksheet = sheet
        self._worksheet_wr = sheet_wr
        self._rows = []
        self._min_row = 0
        self._max_row = None
        self._cols = []
        self._min_col = 0
        self._max_col = None
        self._title = {}
        self._last_xmlname_row = 0
        if sheet != None:
            self.init_sheet()
        self.init_model()
    def __del__(self):
        del self._worksheet
        del self._worksheet_wr

    def init_sheet(self):
        self._xmlname = self.search_by_value('xmlname')

    
    def init_model(self): 
        self._header_model = QStandardItemModel()
        self._xml_name_model = QStandardItemModel()

    ##################################################
    #       Model api
    ##################################################
    def update_model(self):
        self.init_model()
        self.init_sheet()
        if self._xmlname != None:
            for header in self.headers():
                item_header = QHeaderItem(header)
                item_header.setCheckState(Qt.Unchecked)
                item_header.setCheckable(True)
                self._header_model.appendRow(item_header)

    ##################################################
    #       model data operation
    ##################################################
    def search_by_value(self,value):
        for row in self.rows:
            for cell in row:
                if cell.value == value:
                    return Workcell(cell,self._worksheet_wr)
        return None
    def search_header_by_value(self,value):
        for row in self._worksheet.iter_rows(min_col=self.min_col,min_row=self._xmlname.row,max_col=self.max_col,max_row=self._xmlname.row):
            for cell in row:
                if cell.value == value:
                    return Header(cell,self._worksheet_wr)
        return None

    def search_xmlname_by_value(self,value):
        for col in self._worksheet.iter_cols(min_col=self._xmlname.col,min_row=self._xmlname.row+1,max_col=self._xmlname.col,max_row=self.max_row):
            for cell in col:
                if cell.value == value:
                    return XmlName(cell,self._worksheet_wr)
        return None
    def xml_names(self):
        cells = list(self._worksheet.iter_cols(min_col=self._xmlname.col,min_row=self._xmlname.row+1,max_col=self._xmlname.col,max_row=self.max_row).next())
        while cells[-1].value == None:
            cells.pop()
        self._last_xmlname_row = cells[-1].row
        return map(lambda x:XmlName(x,self._worksheet_wr),cells)
    def xml_names_value(self):
        cells = list(self._worksheet.iter_cols(min_col=self._xmlname.col,min_row=self._xmlname.row+1,max_col=self._xmlname.col,max_row=self.max_row).next())
        while cells[-1].value == None:
            cells.pop()
        return map(lambda x:x.value,cells)
    def headers(self):
        cells = list(self._worksheet.iter_rows(min_col=self.min_col,min_row=self._xmlname.row,max_col=self.max_col,max_row=self._xmlname.row).next())
        while cells[-1].value == None:
            cells.pop()
        while cells[0].value == None:
            cells.pop(0)
        return map(lambda x:Header(x,self._worksheet_wr),cells)
    def headers_value(self):
        cells = list(self._worksheet.iter_rows(min_col=self.min_col,min_row=self._xmlname.row,max_col=self.max_col,max_row=self._xmlname.row).next())
        while cells[-1].value == None:
            cells.pop()
        while cells[0].value == None:
            cells.pop(0)
        return map(lambda x:str(x.value) if x.value is not None else '',cells)
    def select_all_headers(self):
        for i in range(self._header_model.rowCount()):
            item = self._header_model.item(i)
            item.setCheckState(Qt.Checked)
    def unselect_all_headers(self):
        for i in range(self._header_model.rowCount()):
            item = self._header_model.item(i)
            item.setCheckState(Qt.Unchecked)
    ##################################################
    #       Worksheet property
    ##################################################
    def cell(self,row,col):
    #return self._worksheet_wr.range(row,col)
        return self._worksheet.cell(row=row,column=col)
    def cell_value(self,row,col):
    #return self._worksheet_wr.range(row,col)
        return self._worksheet.cell(row=row,column=col).value
    def cell_wr(self,row,col):
        return self._worksheet_wr.range(row,col)
    def cell_wr_value(self,row,col):
        return self._worksheet_wr.range(row,col).value
 
    @property
    def xmlname(self):
        return self._xmlname
    @property
    def preview_model(self):
        return self._preview_model
    @property
    def header_model(self):
        return self._header_model
    @property
    def header_list(self):
        if self._xmlname != None:
            return self.headers_value()
        else:
            return []
    @property
    def xml_name_model(self):
        return self._xml_name_model
    @property
    def extended_preview_model(self):
        return self._extended_preview_model
    @property
    def last_xmlname_row(self):
        return self._last_xmlname_row

    ##################################################
    #       Sync
    ##################################################
    def checked_headers(self):
        items = []
        for i in range(self._header_model.rowCount()):
            item = self._header_model.item(i)
            if item.checkState() == Qt.Checked:
                items.append(item.cell)
        return items
                
    ##################################################
    #       Data
    ##################################################
    @property
    def worksheet(self):
        return self._worksheet
    @property
    def rows(self):
        return self._worksheet.rows
    @property
    def cols(self):
        return self._worksheet.columns
    @property
    def max_row(self):
        if self._worksheet.max_row < MAX_ROW:
            return self._worksheet.max_row
        else:
            return MAX_ROW
    @property
    def min_row(self):
        return self._worksheet.min_row
    @property
    def max_col(self):
        return self._worksheet.max_column
    @property
    def min_col(self):
        return self._worksheet.min_column
class QPreviewItem(QStandardItem):
    def __init__(self,cell):
        if cell.value != None:
            super(QPreviewItem,self).__init__(cell.value)
        else:
            super(QPreviewItem,self).__init__('')
        self._cell = cell
    
    @property
    def cell(self):
        return self._cell
    @property
    def value(self):
        return self._cell.value
    @property
    def row(self):
        return self._cell.row
    @property
    def col(self):
        return self._cell.col
    @property
    def col_letter(self):
        return self._cell.col_letter
class QComparisonItem(QStandardItem):
    def __init__(self,cell):
        if cell.value != None:
            super(QComparisonItem,self).__init__(cell.value)
        else:
            super(QComparisonItem,self).__init__('')
        self._cell = cell
    
    @property
    def cell(self):
        return self._cell
    @property
    def value(self):
        return self._cell.value
    @property
    def row(self):
        return self._cell.row
    @property
    def col(self):
        return self._cell.col
    @property
    def col_letter(self):
        return self._cell.col_letter
class QHeaderItem(QStandardItem):
    def __init__(self,cell):
        if cell.value != None:
            super(QHeaderItem,self).__init__(cell.value)
        else:
            super(QHeaderItem,self).__init__('')
        self._cell = cell
    
    def get_item_by_xmlname(self,xmlname):
        return self._cell.get_item_by_xmlname(xmlname)
    
    @property
    def cell(self):
        return self._cell
    @property
    def value(self):
        return self._cell.value
    @property
    def row(self):
        return self._cell.row
    @property
    def col(self):
        return self._cell.col
    @property
    def col_letter(self):
        return self._cell.col_letter
        
 
