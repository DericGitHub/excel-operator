import openpyxl as xl
from openpyxl.cell import Cell
import time
from threading import Timer
from copy import copy
#wb = xl.load_workbook('Swan_M227_CAS_final.xlsx')
#print wb.sheetnames
#print wb.defined_name
#ws = wb['Swan M227']
#print ws.cell(row=5,column=7).value
UP = -1
DOWN = 1
LEFT = -1
RIGHT = 1

class jxl():
    def __init__(self,file_path):
        self.file_path = file_path
        self.workbook = xl.load_workbook(file_path)
        self.autosave_on = False
        self.save_period = None
        result ='''Inited jxl instance.
        file_path = %s
        autosave_on = %s
        save_period = %s
        '''%(self.file_path,self.autosave_on,self.save_period)
        print(result)

    
    ##################################################
    #       save 
    ##################################################
    def save(self):
        self.workbook.save(self.file_path)
    
    def save_as(self,file_path):
        self.workbook.save(file_path)
        
    ##################################################
    #       fetch 
    ##################################################
    ##################################################
    #       modification
    ##################################################
    def append(self,sheet_name,iterable):
        self.workbook[sheet_name].append(iterable)
    def rewrite(self,sheet_name,start_col,iterable):
        self.workbook[sheet_name].rewrite(start_col,iterable)

    def copy_col(self,sheet_name,src,dst):
        pass
    def copy_row(self,sheet_name,src):
        col = next(self.workbook[sheet_name].iter_rows(min_row=src,max_row=src))
        for cell in col:
            #copy_cell = 
            cell_list.append(cell)
        return cell_list
    def cope_row_to(self,sheet_name,src,dst):
        pass 
    def add_col(self,sheet_name,orientation,start_pos,offset):
        ##################################################
        #       version 0.0.1
        #   col width issue
        #   richtext not supported
        ##################################################
        cols = []
        if orientation == RIGHT:
            real_start_pos = start_pos
        elif orientation == LEFT:
            real_start_pos = start_pos + 1
        for col in self.workbook[sheet_name].iter_cols(min_col=real_start_pos,max_col=self.workbook[sheet_name].max_column):
            cols.append(col)
        print(cols)
        for col in reversed(cols):
            col = reversed(col)
            adjust_col_width = True
            for cell in col:
                print("copy %s cell.row = %d,offset = %d"%(cell,cell.row,offset))
                offset_cell = cell.parent.cell(row=cell.row,column=ord(cell.column)-ord('A')+1+offset)
                print("to %s cell.row = %d,offset = %d"%(offset_cell,offset_cell.row,offset))
                if adjust_col_width == True:
                    self.workbook[sheet_name].column_dimensions[offset_cell.column] = self.workbook[sheet_name].column_dimensions[cell.column]
                    print("col %s width:%s"%(offset_cell.column,self.workbook[sheet_name].column_dimensions[offset_cell.column].width))
                    print("col %s width:%s"%(cell.column,self.workbook[sheet_name].column_dimensions[cell.column].width))
                    adjust_col_width = False
                offset_cell.value = cell.value
                offset_cell.alignment = cell.alignment.copy()
                offset_cell.border = cell.border.copy()
                offset_cell.fill = cell.fill.copy()
                offset_cell.font = cell.font.copy()
                offset_cell.number_format = cell.number_format
                offset_cell.protection = cell.protection.copy()
                #offset_cell.style = cell.style

        for col in self.workbook[sheet_name].iter_cols(min_col=real_start_pos,max_col=start_pos+offset-1):
            adjust_col_width = True
            for cell in col:
                if adjust_col_width == True:
                    self.workbook[sheet_name].column_dimensions[cell.column].width = None
                    adjust_col_width = False
                cell.value = ''

               

            
    def del_col(self,sheet_name,orientation,start_pos,offset):
        pass
    def add_row(self,sheet_name,orientation,start_pos,offset):
        ##################################################
        #       version 0.1.0
        #   richtext not supported 
        ##################################################
        rows = []
        if orientation == UP:
            real_start_pos = start_pos
        elif orientation == DOWN:
            real_start_pos = start_pos + 1
        for row in self.workbook[sheet_name].iter_rows(min_row=real_start_pos,max_row=self.workbook[sheet_name].max_row):
            rows.append(row)
        #print(rows)
        for row in reversed(rows):
            row = reversed(row)
            adjust_row_height = True
            for cell in row:
                print("%s cell.row = %d,offset = %d"%(cell,cell.row,offset))
                offset_cell = cell.parent.cell(row=cell.row+offset,column=ord(cell.column)-ord('A')+1)
                print("%s cell.row = %d,offset = %d"%(offset_cell,offset_cell.row,offset))
                if adjust_row_height == True:
                    self.workbook[sheet_name].row_dimensions[offset_cell.row] = self.workbook[sheet_name].row_dimensions[cell.row]
                    adjust_row_height = False
                offset_cell.value = cell.value
                offset_cell.font = cell.font.copy()
                offset_cell.border = cell.border.copy()
                offset_cell.fill = cell.fill.copy()
                offset_cell.number_format = cell.number_format
                offset_cell.protection = cell.protection.copy()
                offset_cell.alignment = cell.alignment.copy()
        for row in self.workbook[sheet_name].iter_rows(min_row=real_start_pos,max_row=start_pos+offset-1):
            adjust_row_height = True
            for cell in row:
                if adjust_row_height == True:
                    self.workbook[sheet_name].row_dimensions[cell.row].ht = None
                    adjust_row_height = False
                cell.value = ''

               

    def del_row(self,sheet_name,orientation,start_pos,offset):
        pass
    def add_sheet(self,sheet_name,orientation,start_pos,offset):
        pass
    def del_sheet(self,sheet_name,orientation,start_pos,offset):
        pass

    ##################################################
    #       lock
    ##################################################
    def lock(self,cells):
        pass
    def unlock(self,cells):
        pass
    
    ##################################################
    #       comparison
    ##################################################


    ##################################################
    #       autosave
    ##################################################
    def autosave(self):
        if self.autosave_on == True:
            pass 

if __name__ == "__main__":
    handler = jxl("faxtel.xlsx")
    #handler.append('Sheet1',next(handler.workbook['Sheet1'].iter_rows(min_row=5,max_row=5)))
    handler.add_col('Sheet1',RIGHT,3,3)
    #handler.add_row('Sheet1',DOWN,3,3)
    handler.save_as('test1.xlsx')
